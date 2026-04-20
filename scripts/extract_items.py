"""Extract all item rows from 'Sourced' starting at row 19, with embedded images."""
import json
from collections import Counter
from pathlib import Path
from openpyxl import load_workbook

SRC = Path(r'C:/Users/AaryanMehta/Downloads/jira/13 April - I5DIA MOM-4 meetings. (Compiled) v0.1.xlsx')
OUT_DIR = Path(r'C:/Users/AaryanMehta/Downloads/jira/derived')
IMG_DIR = OUT_DIR / 'images'
IMG_DIR.mkdir(parents=True, exist_ok=True)

# Row 18 headers (from deep scan):
HEADERS = {
    0: 'sr', 1: 'start_date', 2: 'project_code', 3: 'category', 4: 'date',
    5: 'project', 6: 'image', 7: 'style_no', 8: 'no_of_sketch', 9: 'sketch_changes',
    10: 'cad_changes', 11: 'diamond_quality', 12: '1kt', 13: 'comments',
    14: 'priority', 15: 'status', 16: 'status_update_date', 17: 'milestone_log',
    29: 'apr13_ddlny_comments',
}

wb = load_workbook(SRC, data_only=True)
ws = wb['Sourced']

# --- Extract images (openpyxl keeps them in ws._images) ---
# We also need their anchor row to match image to item row.
try:
    from openpyxl.utils import get_column_letter
    images = getattr(ws, '_images', []) or []
    print(f'Embedded images: {len(images)}', flush=True)

    img_map = {}  # row_number -> saved image filename
    for idx, img in enumerate(images):
        anchor = img.anchor
        try:
            row = anchor._from.row + 1  # openpyxl anchor is 0-indexed
        except Exception:
            row = None
        # dump bytes
        ext = 'png'
        try:
            data = img._data()
        except Exception:
            data = img.ref.read() if hasattr(img, 'ref') else None
        if data is None:
            continue
        fname = f'img_{idx:04d}_r{row}.{ext}'
        (IMG_DIR / fname).write_bytes(data)
        if row is not None:
            img_map.setdefault(row, []).append(fname)
    print(f'Saved {len(list(IMG_DIR.glob("*.png")))} images to {IMG_DIR}', flush=True)
except Exception as e:
    print(f'image extract failed: {e}', flush=True)
    img_map = {}

# --- Extract item rows (19 through last) ---
items = []
status_counts = Counter()
project_counts = Counter()
category_counts = Counter()

for r_idx, row in enumerate(ws.iter_rows(min_row=19, max_row=ws.max_row, values_only=True), start=19):
    item = {'row': r_idx}
    for col_idx, key in HEADERS.items():
        if col_idx < len(row):
            v = row[col_idx]
            if v is not None and str(v).strip():
                item[key] = str(v).strip() if not hasattr(v, 'isoformat') else v.isoformat()
    # skip empty rows
    meaningful = any(k in item for k in ('style_no', 'category', 'status', 'comments'))
    if not meaningful:
        continue
    if r_idx in img_map:
        item['images'] = img_map[r_idx]
    items.append(item)
    if item.get('status'):
        status_counts[item['status']] += 1
    if item.get('project_code'):
        project_counts[item['project_code']] += 1
    if item.get('category'):
        category_counts[item['category']] += 1

(OUT_DIR / 'items.json').write_text(json.dumps(items, indent=2, default=str))
print(f'Extracted {len(items)} items → items.json', flush=True)

summary = {
    'total_items': len(items),
    'images_attached': sum(1 for i in items if 'images' in i),
    'statuses': status_counts.most_common(),
    'project_codes': project_counts.most_common(20),
    'categories_top20': category_counts.most_common(20),
    'unique_categories': len(category_counts),
}
(OUT_DIR / 'items_summary.json').write_text(json.dumps(summary, indent=2, default=str))

print('\n=== STATUSES ===')
for s, c in summary['statuses']:
    print(f'  {c:4d}  {s}')
print('\n=== PROJECT CODES ===')
for p, c in summary['project_codes']:
    print(f'  {c:4d}  {p}')
print('\n=== TOP CATEGORIES ===')
for cat, c in summary['categories_top20'][:15]:
    print(f'  {c:4d}  {cat}')
print(f'\nunique categories: {summary["unique_categories"]}')
