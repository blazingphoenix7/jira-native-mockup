"""Deep-scan the 'Sourced' sheet: find the real header row, extract data + image links."""
import json
from pathlib import Path
from openpyxl import load_workbook

SRC = Path(r'C:/Users/AaryanMehta/Downloads/jira/13 April - I5DIA MOM-4 meetings. (Compiled) v0.1.xlsx')
OUT_DIR = Path(r'C:/Users/AaryanMehta/Downloads/jira/derived')
OUT_DIR.mkdir(parents=True, exist_ok=True)

print('Loading workbook ...', flush=True)
wb = load_workbook(SRC, data_only=True)
ws = wb['Sourced']
print(f'dims: {ws.max_row} rows x {ws.max_column} cols', flush=True)

# Dump the top 15 rows fully so we can see where real headers live.
top_rows = []
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
    top_rows.append({'row': i, 'cells': [str(c)[:60] if c is not None else '' for c in row]})

(OUT_DIR / 'sourced_top15.json').write_text(json.dumps(top_rows, indent=2, default=str))
print(f'Wrote sourced_top15.json', flush=True)

# Print densely so we can eyeball
for r in top_rows:
    nonempty = [(i, c) for i, c in enumerate(r['cells']) if c]
    if nonempty:
        print(f"r{r['row']:2d}: {nonempty[:20]}{' ...' if len(nonempty)>20 else ''}")
