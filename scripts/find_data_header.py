"""Scan rows 14-40 to locate the actual item-level header and sample data rows."""
import json
from pathlib import Path
from openpyxl import load_workbook

SRC = Path(r'C:/Users/AaryanMehta/Downloads/jira/13 April - I5DIA MOM-4 meetings. (Compiled) v0.1.xlsx')
OUT = Path(r'C:/Users/AaryanMehta/Downloads/jira/derived/sourced_rows14_40.json')

wb = load_workbook(SRC, data_only=True)
ws = wb['Sourced']

dump = []
for i, row in enumerate(ws.iter_rows(min_row=14, max_row=40, values_only=True), start=14):
    cells = [str(c)[:60] if c is not None else '' for c in row]
    dump.append({'row': i, 'cells': cells})
    nonempty = [(idx, c) for idx, c in enumerate(cells) if c]
    if nonempty:
        print(f"r{i:2d} ({len(nonempty)} non-empty): {nonempty[:25]}")

OUT.write_text(json.dumps(dump, indent=2, default=str))
print(f'Wrote {OUT}')
