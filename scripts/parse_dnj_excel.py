"""Inventory the DNJ weekly report: sheets, dims, headers, sample rows, image count."""
import json
from pathlib import Path
from openpyxl import load_workbook

SRC = Path(r'C:/Users/AaryanMehta/Downloads/jira/13 April - I5DIA MOM-4 meetings. (Compiled) v0.1.xlsx')
OUT = Path(r'C:/Users/AaryanMehta/Downloads/jira/derived/dnj_excel_inventory.json')
OUT.parent.mkdir(parents=True, exist_ok=True)

print(f'Loading {SRC.name} ({SRC.stat().st_size/1e6:.1f} MB) ...', flush=True)
wb = load_workbook(SRC, read_only=True, data_only=True)
print(f'Sheets: {wb.sheetnames}', flush=True)

inventory = {'file': SRC.name, 'size_mb': round(SRC.stat().st_size/1e6, 1), 'sheets': []}

for name in wb.sheetnames:
    ws = wb[name]
    print(f'  scanning "{name}" ...', flush=True)
    rows = list(ws.iter_rows(values_only=True))
    max_col = max((len(r) for r in rows), default=0)
    headers = []
    if rows:
        for r in rows[:6]:
            if any(c not in (None, '') for c in r):
                headers = [str(c) if c is not None else '' for c in r]
                break
    samples = []
    for r in rows[1:8]:
        if any(c not in (None, '') for c in r):
            samples.append([str(c)[:80] if c is not None else '' for c in r[:max_col]])
        if len(samples) >= 3:
            break
    inventory['sheets'].append({
        'name': name,
        'rows': len(rows),
        'cols': max_col,
        'header_row_candidate': headers,
        'sample_rows': samples,
    })

# The file is 99MB — likely embedded images. Count them via openpyxl's image storage.
# (read_only mode doesn't expose images; reopen normally just for image count, sheet by sheet.)
print('Re-opening (non-read-only) to count embedded images ...', flush=True)
wb2 = load_workbook(SRC, data_only=True)
for s in inventory['sheets']:
    ws = wb2[s['name']]
    try:
        s['embedded_images'] = len(getattr(ws, '_images', []) or [])
    except Exception as e:
        s['embedded_images'] = f'err: {e}'

OUT.write_text(json.dumps(inventory, indent=2, default=str))
print(f'Wrote {OUT}', flush=True)

# Console summary
for s in inventory['sheets']:
    print(f"  - {s['name']}: {s['rows']} rows x {s['cols']} cols, images={s.get('embedded_images')}")
